import pandas as pd
import xlsxwriter
import os
from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill
import json

PARALLEL_NUM=6

def read_maze():
	df=pd.read_excel("maze"+str(PARALLEL_NUM)+".xlsx",header=None)
	df=df.applymap(lambda x:json.loads(x))
	df=df.iloc[::-1]
	# rearrange dataframe rows
	return df.values.tolist()

def update_xl(maze):
	wb=Workbook()
	maze_ws=wb.active
	#print(dir(maze_ws))
	maze_ws.sheet_view.zoomScale=22
	maze_ws.column_dimension=0.1
	maze_ws.row_dimension=0.1
	color_ls=['00000000','0000FF00','00FFFF00','FFFFFFFF','00FF0000']
	for i in range(250):
	# i is xl column==x
		for j in range(250):
		# j is xl row==y
			maze_ws[xl_index([2*i,2*j])]=str((i,249-j))+str(maze[249-j][i])
			maze_ws[xl_index([2*i,2*j+1])].fill=PatternFill("solid",fgColor=color_ls[maze[249-j][i]["d"]])
			maze_ws[xl_index([2*i+1,2*j])].fill=PatternFill("solid",fgColor=color_ls[maze[249-j][i]["r"]])
			#maze_ws[xl_index([2*i+1,2*j+1])].fill=PatternFill("solid",fgColor=color_ls[3])
			
	wb.save("color_maze"+str(PARALLEL_NUM)+".xlsx")
	

def xl_index(loc):
	# input: x,y
	# output letter rep in excel , 250-row
	let=""
	if int(loc[0]/26)>0:
		let+=chr(int(loc[0]/26)+64)
	let+=chr(loc[0]%26+65)
	return let+str(loc[1]+1)
	
maze=read_maze()
update_xl(maze)
os.system('play -nq -t alsa synth {} sine {}'.format(1,440))
