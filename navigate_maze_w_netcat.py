import pandas as pd
import time
import sys
import socket
import xlsxwriter
from openpyxl import load_workbook,Workbook
import json
import os
from sympy import symbols, Eq, solve

# change for each parallel run: new_maze#, new_dist&hints#

MAZE_SIZE=250
PARALLEL_NUM=5
MAX_RADIUS=70

def initiate_maze():
	wb=Workbook()
	maze_ws=wb.active
	for i in range(MAZE_SIZE):
		for j in range(MAZE_SIZE):
			maze_ws[xl_index([i,j])]=str('{"l":0,"u":0,"d":0,"r":0}') 
	wb.save("maze"+str(PARALLEL_NUM)+".xlsx")

def update_xl(maze):
	wb=Workbook()
	maze_ws=wb.active
	for i in range(MAZE_SIZE):
		for j in range(MAZE_SIZE):
			maze_ws[xl_index([i,j])]=str(maze[i][j]).replace("'","\t").replace('"',"'").replace("\t",'"')  
	print("Updating maze"+str(PARALLEL_NUM)+".xlsx")
	wb.save("maze"+str(PARALLEL_NUM)+".xlsx")

def read_maze():
	df=pd.read_excel("maze"+str(PARALLEL_NUM)+".xlsx",header=None)
	df=df.applymap(lambda x:json.loads(x))
	return df.values.tolist()

"""
in xl - {"l": 0,"u": 0,"d": 0,"r": 0}
in maze - {'l':0,'u':0,'d':0,'r':0}
"""


def xl_index(loc):
	let=""
	if int(loc[0]/26)>0:
		let+=chr(int(loc[0]/26)+64)
	let+=chr(loc[0]%26+65)
	return let+str(250-loc[1])
"""
maze direction values
-1 blocked 
0 untraveled path
1 traveled, didnt complete
2 traveled, got to all dead ends
3 trap in the dead end ahead
"""
def next_loc(loc,direc):
	"""
	input: index pair of [row (int), column (int)] and direction (str)
	output: index pair of [row (int), column (int)] after proceeding to 		the right direction
	"""
	if direc=="u":
		return [(loc[0])%250,(loc[1]+1)%250]
	elif direc=="r":
		return [(loc[0]+1)%250,(loc[1])%250]
	elif direc=="d":
		return [(loc[0])%250,(loc[1]-1)%250]
	elif direc=="l":
		return [(loc[0]-1)%250,(loc[1])%250]

def blocked(maze,loc,direc):
	#print(f"{loc} blocked in {direc} direction, stayed in {loc}")
	maze[loc[0]][loc[1]][direc]=-1
	maze[next_loc(loc,direc)[0]][next_loc(loc,direc)[1]][opp(direc)]=-1


def travel(maze,loc,direc):
	#print(f"traveling from {loc} to {next_loc(loc,direc)} succesfully")
	next=next_loc(loc,direc)
	if check_leaving_dead_end(maze[loc[0]][loc[1]],direc):
		maze[loc[0]][loc[1]][direc]=2
		maze[next[0]][next[1]][opp(direc)]=2
		#print(f"^^^^^^^^^^^^^^ leaving a dead end")
	else:
		maze[loc[0]][loc[1]][direc]=1
		maze[next[0]][next[1]][opp(direc)]=1
		#print(f"^^^^^^^^^^^^^^ didn't complete behind me")
	return next

def insert_blocks(maze,loc,info):
	for direc,i in zip(["l","r","u","d"],[2,7,12,17]):
		if info[i]=="0":
			blocked(maze,loc,direc)
	
	
def show_area(maze,loc,target_loc,direc):
	print("\nSurrounding Area:\n")
	buf=2
	dic_string=""
	rows=maze[(loc[0]-buf)%250:(loc[0]+buf+1)%250]
	left_to_target=dir_to_target(maze,direc,target_loc,loc)
	next_preffered_dir=["turn right","turn left"][left_to_target]
	next_turns=["turn counter clockwise","turn clockwise"][left_to_target]
	if direc=="r":
		if left_to_target==1:
			dic_string+=("\t\t\t\t\t\t^^^^^^>>\n")
	elif direc=="u":
		if left_to_target==1:
			dic_string+=("^\n^\n<\n<\n")
		else:
			dic_string+=("\t\t\t\t\t\t\t\t\t\t\t\t^\n\t\t\t\t\t\t\t\t\t\t\t\t^\n\t\t\t\t\t\t\t\t\t\t\t\t>\n\t\t\t\t\t\t\t\t\t\t\t\t>\n")
	elif direc=="l":
		if left_to_target==0:
			dic_string+=("\t\t\t\t\t<<^^^^^^\n")
	elif direc=="d":
		if left_to_target==1:
			dic_string+=("\t\t\t\t\t\t\t\t\t\t\t\t>\n\t\t\t\t\t\t\t\t\t\t\t\t>\n")
		else:
			dic_string+=("<\n<\n")
			
	if (direc,next_preffered_dir) in [("r","turn left"),("l","turn right")]:
		dic_string+= ("\t\t\t\t\t\t^^^^\n")
	
	for j in range(buf*2+1):
		for i in range(buf*2+1):
			dic_string+=str(maze[(loc[0]-buf+i)%250][loc[1]+buf-j])+","
		dic_string+="\n"
	
	if direc=="r" and left_to_target==0:
			dic_string+=("\t\t\t\t\t<<VVVVVV\n")
			
	elif direc=="u":
		if left_to_target==1:
			dic_string+=("<\n<\n")
		else:
			dic_string+=("\t\t\t\t\t\t\t\t\t\t\t\t>\n\t\t\t\t\t\t\t\t\t\t\t\t>\n")
	elif direc=="l" and left_to_target==1:
			dic_string+=("\t\t\t\t\t<<VVVVVV\n")
	elif direc=="d":
		if left_to_target==1:
			dic_string+=("\n\t\t\t\t\t\t\t\t\t\t\t\t>\n\t\t\t\t\t\t\t\t\t\t\t\t>\t\t\t\t\t\t\t\t\t\t\t\tV\n\t\t\t\t\t\t\t\t\t\t\t\tV\n")
		else:
			dic_string+="<\n<\nV\nV\n"
	print(dic_string)
			
	
			
			
def check_leaving_dead_end(dir_dic,direc):
	"""
	checks to see if moving from location in direc leaves loc a dead end (only blocks / dead 		ends)
	"""
	for key in dir_dic:
		if str(key)!=direc:
			if dir_dic[key] in [0,1]:
				return 0
	return 1 

def opp(direc):
	"""
	input: direction (str)
	output: opposite direction (str)
	"""
	if direc=="u":
		return "d"
	elif direc=="r":
		return "l"
	elif direc=="d":
		return "u"
	elif direc=="l":
		return "r"


def next_dir_to_solution(maze,dir_dic,direc,turnbacks,target_loc,loc,closest):
	# dir_dic is paths from location, direc is last direction
	# prefer dir to target, if not then straight ahead, then opossite side, turn back if dead end
	# if we are going further from target, 
	way_to_target=dir_to_target(maze,direc,target_loc,loc)
	turn_to_target,turn_away_from_target=[counter_cw,cw][::[-1,1][way_to_target]]
	
	if distance(target_loc,loc)>MAX_RADIUS and closest<25:
		print ("Had to turn back to the target")
		#os.system('play -nq -t alsa synth {} sine {}'.format(1,440))
		time.sleep(2)
		
		return opp(direc),turnbacks+1
	
	for pref in range(2):
		next=turn_to_target(direc)
		while (next!=opp(direc)):
			#print (f"To turn into the target we need to turn our direction {turn_to_target.__name__}, our options are: {dir_dic}, so we went to {next}")
			if dir_dic[next]==pref:
				return next,0
			else:
				next=turn_away_from_target(next)
	return opp(direc),turnbacks+1

def distance(target_loc,loc):
	diff_y=(target_loc[0]-loc[0]+100)%250-100
	# pos if needs to move up, neg if needs to move down
	diff_x=(target_loc[1]-loc[1]+100)%250-100
	# pos if needs to move right, neg if needs to move left
	return int((diff_x**2+diff_y**2)**0.5)

def going_far(target_loc,loc,direc,):
	diff_y=(target_loc[1]-loc[1]+100)%250-100
	# pos if needs to move up, neg if needs to move down
	diff_x=(target_loc[0]-loc[0]+100)%250-100
	# pos if needs to move right, neg if needs to move left
	
	far_dirs=[]
	if diff_y>0:
		far_dirs.append("d")
	elif diff_y<0:
		far_dirs.append("u")
	if diff_x>0:
		far_dirs.append("l")
	elif diff_x<0:
		far_dirs.append("r")
	if direc in far_dirs:
		#print("going far from {target_loc}")
		return 1
	else:
		#print("going closer to {target_loc}")
		return 0  

def dir_to_target(maze,direc,target_loc,loc):
	diff_y=(target_loc[1]-loc[1]+100)%250-100
	# pos if needs to move up, neg if needs to move down
	diff_x=(target_loc[0]-loc[0]+100)%250-100
	# pos if needs to move right, neg if needs to move left
	#print(f"diff_x: {diff_x}, diff_y: {diff_y}")

	crossers=0

	# count crossers in the row adjustment
	if diff_x!=0:
		i=int(diff_x/abs(diff_x))
		while abs(i)<abs(diff_x):
			if maze[(loc[0]+i)%250][loc[1]]["u"]==1:
				while abs(i)<abs(diff_x) and (maze[(loc[0]+i)%250][loc[1]]["r" if diff_x>0 else "l"]==1 or maze[(loc[0]+i)%250][loc[1]]["d"]==1):
					if maze[(loc[0]+i)%250][loc[1]]["d"]==1:
						crossers+=1
						break
					else:
						i+=int(diff_x/abs(diff_x))
			elif maze[(loc[0]+i)%250][loc[1]]["d"]==1:
				while abs(i)<abs(diff_x) and (maze[(loc[0]+i)%250][loc[1]]["r" if diff_x>0 else "l"]==1 or maze[(loc[0]+i)%250][loc[1]]["u"]==1):
					if maze[(loc[0]+i)%250][loc[1]]["u"]==1:
						crossers+=1
						break
					else:
						i+=int(diff_x/abs(diff_x))

			i+=int(diff_x/abs(diff_x))
	if diff_y!=0:
		i=int(diff_y/abs(diff_y))
		while abs(i)<abs(diff_y):
			if maze[target_loc[0]][(loc[1]+i)%250]["l"]==1:
				while abs(i)<abs(diff_y) and (maze[target_loc[0]][(loc[1]+i)%250]["u" if diff_y>0 else "d"]==1 or maze[target_loc[0]][(loc[1]+i)%250]["r"]==1):
					if maze[target_loc[0]][(loc[1]+i)%250]["r"]==1:
						crossers+=1
						break
					else:
						i+=int(diff_y/abs(diff_y))
			elif maze[target_loc[0]][(loc[1]+i)%250]["r"]==1:
				while abs(i)<abs(diff_y) and (maze[target_loc[0]][(loc[1]+i)%250]["u" if diff_y>0 else "d"]==1 or maze[target_loc[0]][(loc[1]+i)%250]["l"]==1):
					if maze[target_loc[0]][(loc[1]+i)%250]["l"]==1:
						crossers+=1
						break
					else:
						i+=int(diff_y/abs(diff_y))
			i+=int(diff_y/abs(diff_y))
			
	#print(f"num of crossers:{crossers}")
		
	if direc=="u":
		if diff_x<0:
			#print("Need to be left oriented to get to target")
			return (crossers+1)%2
		else:
			#print("Need to be right oriented to get to target")
			return (crossers)%2
	if direc=="r":
		if diff_y>0:
			#print("Need to be left oriented to get to target")
			return (crossers+1)%2
		else:
			#print("Need to be right oriented to get to target")
			return (crossers)%2
	if direc=="d":
		if diff_x>0:
			#print("Need to be left oriented to get to target")
			return (crossers+1)%2
		else:
			#print("Need to be right oriented to get to target")
			return (crossers)%2
	if direc=="l":
		if diff_y<0:
			#print("Need to be left oriented to get to target")
			return (crossers+1)%2
		else:
			#print("Need to be right oriented to get to target")
			return (crossers)%2
			
			
			
def next_dir_free_roam(dir_dic,direc,turnbacks):
	# dir_dic is paths from location, direc is last direction
	# roads untravelled (=0), travelled and not dead_end (=1, preferably not to last location), 	prefer nearest cw to opposite
	for pref in range(2):
		next=counter_cw(direc)
		while (next!=opp(direc)):
			if dir_dic[next]==pref:
				return next,0
			else:
				next=cw(next)
	return opp(direc),turnbacks+1

def cw(direc):
	return {"u":"r","r":"d","d":"l","l":"u"}[direc]

def counter_cw(direc):
	return {"u":"l","r":"u","d":"r","l":"d"}[direc]

def check_boundries(loc,dir_dic,direc,turnbacks):
		if loc[0]<0 and direc=="u":
			print("VVVVVVVV  HIT TOP, GO DOWN!!  VVVVVVVV")
			return "d",turnbacks+1
		elif loc[0]>=250 and direc=="d":
			print("^^^^^^^^  HIT BOTTOM, GO UP!!  ^^^^^^^^")
			return "u",turnbacks+1
		elif loc[1]<0 and direc=="l":
			print(">>>>>>>>  HIT LEFT, GO RIGHT!!  >>>>>>>>")
			return "r",turnbacks+1
		elif loc[1]>=250 and direc=="r":
			print("<<<<<<<<  HIT RIGHT, GO LEFT!!  <<<<<<<<")
			return "l",turnbacks+1
		else:
			return direc,turnbacks
"""
def correct_250(loc1,loc):
	if abs(loc[0]-loc1[0])<abs(abs(loc[0]-loc1[0])-250)
		x=loc1[0]
	else:
		x=
"""

def connect():
	maze=read_maze()
	clear=lambda: os.system('clear')
	clear()
	num_of_steps=0
	num_of_turnbacks=0
	dis_and_loc=[]
	closest=50
	with open("dist&hints"+str(PARALLEL_NUM)+".txt","w") as f:
		f.write("")
	direc_display={'u':' ^^^ ','r':'    >','d':'  V  ','l':'<    '}
	buf=200
	s=socket.socket(socket.AF_INET, socket.SOCK_STREAM)
	s.connect(('maze.csa-challenge.com', 80))
	text=""
	for i in range(12):
		text+=str(s.recv(buf))

	loc_text=text.split("(")[1].split(")")[0].split(",")
	loc=[int(loc_text[0]),int(loc_text[1])]

	solution=(-1,-1)
	print(loc)
	direc="u"
	cmd="n"
	while direc!="q":
		#clear()
		s.send(bytes("i",'utf-8'))
		reply_to_direc=s.recv(buf).decode("utf-8")
		insert_blocks(maze,loc,reply_to_direc)
		#print(f"{num_of_steps}-{loc}-last_direc: ",direc_display[direc])
		
		if solution==(-1,-1):
		# didn't find the treasure location yet
			direc,num_of_turnbacks=next_dir_free_roam(maze[loc[0]][loc[1]],direc,num_of_turnbacks)
		else:
		# found treasure location!
			cur_dist=distance(solution,loc) 
			if closest>cur_dist:
				closest=cur_dist
			direc,num_of_turnbacks=next_dir_to_solution(maze,maze[loc[0]][loc[1]],direc,num_of_turnbacks,solution,loc,closest)
			
		text=s.recv(buf).decode("utf-8") # == "> What is your command?"
		
		
		if loc==solution:
			print("Entered Solution Location!")
			update_xl(maze)
			
			s.send(bytes('g','utf-8'))
			dist=s.recv(buf).decode("utf-8")
			print(f"dist in solution: {dist}")
			text=s.recv(buf).decode("utf-8")
			print(f"text in solution: {text}")
			
			s.send(bytes('i','utf-8'))
			info=s.recv(buf).decode("utf-8")
			print(f"info in solution: {info}")
			text=s.recv(buf).decode("utf-8")
			print(f"text in solution: {text}")
			
			for i in range(10):
				os.system('play -nq -t alsa synth {} sine {}'.format(0.3,440))
			cmd=input(f"what is next command? (p - recv more/ i - info / g - distance) ")
			while cmd=="p":
				text=s.recv(buf).decode("utf-8") # reply to solution in correct location
				cmd=input(f"reply is: {text}, what is next command? (p/i/g) ")
			

		s.send(bytes('g','utf-8'))
		dist=s.recv(buf).decode("utf-8")
		text=s.recv(buf).decode("utf-8") # == "> What is your command?"
		if dist[:5]!="far f":
			if solution!=(-1,-1):
				print(f"{loc}-{cur_dist}-{len(dis_and_loc)}-{solution}")
			else:
				print(f"{loc}-{len(dis_and_loc)}")
			dis=int(dist.split()[-1].replace("\n",""))
			dis_and_loc.append((dis,loc))
			s.send(bytes('h','utf-8'))
			hint=s.recv(buf).decode("utf-8")
			text=s.recv(buf).decode("utf-8") # == "> What is your command?"
			if hint[:5] in ["I wis","I bel","Reall","Don't","Don't"]:
				hint="irrelevent "
			with open("dist&hints"+str(PARALLEL_NUM)+".txt","a") as f:
				f.write(f"steps:{num_of_steps} loc:{loc} d:{dist[:-1]}\n")
		else:
			if solution!=(-1,-1):
				print(f"{loc}-{cur_dist}-{len(dis_and_loc)}-{solution}")
		
		if len(dis_and_loc)%100!=3:	
				
			# send direction command to socket
			
			s.send(bytes(str(direc),'utf-8'))
			reply_to_direc=s.recv(buf).decode("utf-8")
			
			if reply_to_direc[0]=="0":
				blocked(maze,loc,direc)
			elif reply_to_direc[0]=="1":
				loc=travel(maze,loc,direc)
			else:
				emer=input(f"reply_to_direc: {reply_to_direc}, text: {text},  press enter to proceed")
				if emer=="c":
					continue
			text=s.recv(buf).decode("utf-8") # == "> What is your command?"
			
			if text[:5]!="> Wha":
				input(f"text: {text} press enter to proceed")
			num_of_steps+=1
			if num_of_steps%5000==0:
				update_xl(maze)
			if num_of_turnbacks==2:
				update_xl(maze)
				if input("Done with this path, enter q to stop. ")=="q":
					print("Stopping")
					direc="q"
		else:
			print("Got close enough (again?)")
			update_xl(maze)
			if len(dis_and_loc)==3:
				eqs=[""]*3
				d1,loc1=dis_and_loc[0]
				x1,y1=loc1
				d2,loc2=dis_and_loc[1]
				x2,y2=loc2
				d3,loc3=dis_and_loc[-1]
				x3,y3=loc3
				x,y=symbols('x y')
				D=[d1,d2,d3]
				X=[x1,x2,x3]
				Y=[y1,y2,y3]
				for i in range(3):
					if X[i]<50:
						X[i]+=250
					if Y[i]<50:
						Y[i]+=250
					eqs[i]=((x-X[i])**2+(y-Y[i])**2-D[i])
				sol_dic=solve(eqs,(x,y))
				solution=sol_dic[0]
				solution=[var if var<250 else var-250 for var in solution]
			else:
				print(maze[loc[0]][loc[1]])
			print(f"the Nuaglamir is in {str(solution)}")
			if distance(solution,loc)<5:
				for i in range(5):
					os.system('play -nq -t alsa synth {} sine {}'.format(0.3,440))
				cmd=input("We are close. Do you want to take over? (y/n/q) ")
			else:
				cmd="n"
				print("We are close. Didn't let you take over ")
			if cmd=="q":
				print("Stopping")
				direc="q"
			elif cmd=="y":
				show_area(maze,loc,solution,direc)
				while cmd in "yiludrghspa":
					
					cmd=input("\nWhat is the command?\ni - info\np - recv\ng - dis\nh - hint\ns - solution\nn - return to auto\na - alter target\nDid you get '> What is your command yet?' ")
					if cmd=="a":
						solution=list(input("write correct target : '(x0,y0): '"))
					elif cmd=="q":
						direc="q"
					elif cmd!="n":
						if cmd!="p":
							if cmd in "ludr":
								direc=cmd
							s.send(bytes(cmd,'utf-8'))
						reply_to_direc=s.recv(buf).decode("utf-8")
						if reply_to_direc[0]=="1":
							loc=travel(maze,loc,direc)
							print(f"Traveled to: {loc}, Target is: {solution}")
						print(reply_to_direc)
						show_area(maze,loc,solution,direc)
						if cmd[0]=="(":
							cmd="y"
			else:
				cmd="n"
					
				

initiate_maze()
connect()
